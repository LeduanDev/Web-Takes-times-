from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
import pytz
from ..models import Actividad, Tipo, Estudio, Area, Maquina, Vuelta
from django.http import  HttpResponse
from openpyxl.utils import get_column_letter
from django.shortcuts import get_object_or_404
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook


@login_required
def exportar_excel(request):
    MAX_COLUMN_WIDTH = 20 # Ajusta este valor según tus necesidades

    estudios_seleccionados = request.GET.get('estudios_seleccionados', '').split(',')
    estudios_seleccionados = list(map(int, estudios_seleccionados))

    # Pre-cargar los detalles de estudio
    detalles_estudio = {}
    for estudio_id in estudios_seleccionados:
        estudio = get_object_or_404(Estudio, id_estudio=estudio_id)
        detalles_estudio[estudio_id] = {
            "nombre": estudio.nombre,
            "descripcion": estudio.descripcion,
            "maquina": estudio.maquina.nombre,
            "area": estudio.area.nombre,
            "fecha_inicio": estudio.tiempo_inicio.astimezone(pytz.timezone('America/Bogota')).strftime("%Y-%m-%d %H:%M:%S") if estudio.tiempo_inicio else None,
            "fecha_fin": estudio.tiempo_fin.astimezone(pytz.timezone('America/Bogota')).strftime("%Y-%m-%d %H:%M:%S") if estudio.tiempo_fin else None,
            "duracion": estudio.duracion,
        }

    # Crear el libro de trabajo
    wb = Workbook()

    # Configurar la respuesta HTTP para el archivo Excel
    response = HttpResponse(content_type="application/ms-excel")
    primer_estudio_id = estudios_seleccionados[0]
    primer_estudio = detalles_estudio[primer_estudio_id]
    response["Content-Disposition"] = 'attachment; filename="{}.xlsx"'.format(f"Estudio: {primer_estudio['nombre']}")


    # Configuración de estilo y centrado
    font_barlow = Font(name="Barlow", size=12)
    alignment_center = Alignment(horizontal="center", vertical="center")

    for counter, estudio_id in enumerate(estudios_seleccionados, start=1):
        estudio = get_object_or_404(Estudio, id_estudio=estudio_id)

        # Crear hoja de trabajo para el estudio y las vueltas de las actividades
        ws_estudio = wb.create_sheet(title=estudio.nombre)  # Utilizar el nombre del estudio como título de la hoja

        # Agregar encabezados personalizados para los detalles del estudio
        ws_estudio["A1"] = "Nombre del Estudio:"
        ws_estudio["B1"] = estudio.nombre
        ws_estudio["A1"].font = Font(bold=True)
        ws_estudio["B1"].font = font_barlow
        ws_estudio["B1"].alignment = alignment_center

        # Detalles del estudio
        detalles_estudio = {
            "#": counter,
            "Nombre de estudio": estudio.nombre,
            "Descripción": estudio.descripcion,
            "Máquina": estudio.maquina.nombre,
            "Área": estudio.area.nombre,
            "Fecha de inicio": estudio.tiempo_inicio.astimezone(pytz.timezone('America/Bogota')).strftime("%Y-%m-%d %H:%M:%S") if estudio.tiempo_inicio else None,
            "Fecha de fin": estudio.tiempo_fin.astimezone(pytz.timezone('America/Bogota')).strftime("%Y-%m-%d %H:%M:%S") if estudio.tiempo_fin else None,
            "Tiempo total": estudio.duracion,
        }

        # Agregar los detalles del estudio a la hoja
        ws_estudio.append(["#", "Nombre de estudio", "Descripción", "Máquina", "Área", "Fecha de inicio", "Fecha de fin", "Tiempo total"])
        ws_estudio.append(list(detalles_estudio.values()))

        # Agregar filas vacías como separación visual
        numero_de_filas_vacias = 1
        for _ in range(numero_de_filas_vacias):
            ws_estudio.append([])

        # Agregar los encabezados de las vueltas de actividades
        encabezados_actividades = [ 
            "#",
            "Actividad",
            "Tipo",
            "Tiempo real(Seg)",
            "Tiempo real(Min)",
            "Tiempo transcurrido",
            "Comentarios",
    
        ]

        # Aplicar formato negrita, tipo de letra y centrado para los encabezados de ambas tablas
        for row in ws_estudio.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(encabezados_actividades)):
            for cell in row:
                cell.font = Font(bold=True, name="Barlow", size=12)
                cell.alignment = alignment_center

        # Agregar los encabezados de las vueltas de actividades
        ws_estudio.append(encabezados_actividades)

        # Aplicar formato negrita, tipo de letra y centrado para los encabezados de las vueltas de actividades
        for row in ws_estudio.iter_rows(min_row=3, max_row=3, min_col=1, max_col=ws_estudio.max_column):
            for cell in row:
                cell.font = Font(bold=True, name="Barlow", size=12)
                cell.alignment = alignment_center

        # Obtener vueltas de actividades y ordenar por fecha de registro
        vueltas_actividades = []

        for actividad in estudio.actividad_set.all():
            vueltas = actividad.vueltas.all().order_by('fecha_registro')
            vueltas_actividades.extend(list(vueltas))

        # Ordenar todas las vueltas por fecha de registro
        vueltas_actividades.sort(key=lambda x: x.fecha_registro)

        # Agregar las vueltas de actividades a la hoja de cálculo
        for counter, vuelta in enumerate(vueltas_actividades, start=1):
            tiempo_real_seconds = vuelta.tiempo_real.total_seconds() if vuelta.tiempo_real else None
            tiempo_real_minutes = round(vuelta.tiempo_real.total_seconds() / 60, 2) if vuelta.tiempo_real else None
            tiempo_total_minutes = round(vuelta.tiempo_total.total_seconds() / 60, 2)

            # Si es el primer registro, mostrar el tiempo total como tiempo real
            if counter == 1:
                tiempo_real_minutes = tiempo_total_minutes
                tiempo_real_seconds = vuelta.tiempo_total.total_seconds() if vuelta.tiempo_total else None
            
            comentarios_vuelta = ', '.join(vuelta.comentarios_vuelta()) if vuelta.comentarios_vuelta() else None

            ws_estudio.append([
                counter,
                vuelta.actividad.nombre,
                vuelta.actividad.tipo.nombre,
                tiempo_real_seconds,
                tiempo_real_minutes,
                vuelta.tiempo_total,
                comentarios_vuelta,
                
            ])
        # Ajustar el ancho de las columnas después de agregar todas las vueltas
        for col_num, value in enumerate(ws_estudio["1"], 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws_estudio[col_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_estudio.column_dimensions[col_letter].width = min(adjusted_width, MAX_COLUMN_WIDTH)


        # Aplicar estilos y centrar solo a los encabezados de las vueltas de actividades
        for row in ws_estudio.iter_rows(min_row=2, max_row=2, min_col=1, max_col=ws_estudio.max_column):
            for cell in row:
                cell.font = Font(bold=True, name="Barlow", size=12)
                cell.alignment = alignment_center

        # Aplicar centrado y estilo para el resto de la hoja de cálculo
        for row in ws_estudio.iter_rows(min_row=3, max_row=ws_estudio.max_row, min_col=1, max_col=ws_estudio.max_column):
            for cell in row:
                cell.font = Font(name="Barlow", size=12)
                cell.alignment = alignment_center

    del wb["Sheet"]

    # Guardar el archivo después de agregar las vueltas
    wb.save(response)

    return response

